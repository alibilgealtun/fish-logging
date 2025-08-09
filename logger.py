from __future__ import annotations

from datetime import datetime
from pathlib import Path
from threading import Lock
from typing import Optional

from openpyxl import Workbook, load_workbook


class ExcelLogger:
    def __init__(self, file_path: Optional[str] = None) -> None:
        self.file_path = Path(file_path) if file_path else Path("logs.xlsx")
        self.lock = Lock()
        self._ensure_workbook()

    def _ensure_workbook(self) -> None:
        with self.lock:
            if not self.file_path.exists():
                wb = Workbook()
                ws = wb.active
                ws.title = "Logs"
                ws.append(["Date", "Time", "Species", "Length (cm)", "Confidence"])
                wb.save(self.file_path)

    def log_entry(self, species: str, length_cm: float, confidence: float) -> None:
        now = datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M:%S")
        with self.lock:
            wb = load_workbook(self.file_path)
            ws = wb.active
            ws.append([date_str, time_str, species, float(length_cm), float(confidence)])
            wb.save(self.file_path)

    def cancel_last(self) -> bool:
        with self.lock:
            wb = load_workbook(self.file_path)
            ws = wb.active
            if ws.max_row <= 1:
                return False
            ws.delete_rows(ws.max_row, 1)
            wb.save(self.file_path)
            return True
