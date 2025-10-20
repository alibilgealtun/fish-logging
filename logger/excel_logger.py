from __future__ import annotations

from datetime import datetime
from pathlib import Path
from threading import Lock
from typing import Optional, List

from core.exceptions import LoggingError


class ExcelLogger:
    def __init__(self, file_path: Optional[str] = None) -> None:
        default_path = Path("logs/hauls/") / "logs.xlsx"
        self.file_path = Path(file_path).absolute() if file_path else default_path.absolute()
        self.lock = Lock()

    def _ensure_workbook(self) -> None:
        """Ensure workbook exists and has correct schema."""
        from openpyxl import Workbook  # type: ignore

        with self.lock:
            if not self.file_path.exists():
                self._create_new_workbook()
            else:
                self._migrate_workbook_schema()

    def _create_new_workbook(self) -> None:
        """Create a new workbook with the current schema.

        Raises:
            LoggingError: If workbook creation fails
        """
        from openpyxl import Workbook  # type: ignore

        self.file_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Logs"
            # Current header includes Station ID
            ws.append(["Date", "Time", "Boat", "Station ID", "Species", "Length (cm)", "Confidence"])
            wb.save(self.file_path)
        except Exception as e:
            raise LoggingError(f"Failed to create workbook: {e}") from e

    def _migrate_workbook_schema(self) -> None:
        """Migrate existing workbook to include Station ID column if missing.

        This is a best-effort migration - failures are logged but don't block logging.
        """
        try:
            from openpyxl import load_workbook  # type: ignore

            wb = load_workbook(self.file_path)
            ws = wb.active
            header = self._read_header(ws)

            if header and "Station ID" not in header:
                # Insert Station ID column after "Boat"
                try:
                    boat_idx0 = header.index("Boat")  # 0-based
                    # openpyxl is 1-based; +1 for position after boat, +1 to convert to 1-based
                    insert_at = boat_idx0 + 2
                except ValueError:
                    insert_at = 4  # fallback to column D

                ws.insert_cols(insert_at, amount=1)
                ws.cell(row=1, column=insert_at).value = "Station ID"
                wb.save(self.file_path)
        except Exception:
            # Best-effort migration; ignore failures to avoid blocking logging
            pass

    def _read_header(self, ws) -> List[str]:
        return [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    def log_entry(self, species: str, length_cm: float, confidence: float, boat: str = "", station_id: str = "") -> None:
        try:
            self._ensure_workbook()
            from openpyxl import load_workbook  # type: ignore

            now = datetime.now()
            date_str = now.strftime("%Y-%m-%d")
            time_str = now.strftime("%H:%M:%S")
            with self.lock:
                wb = load_workbook(self.file_path)
                ws = wb.active
                header = self._read_header(ws)
                if "Station ID" in (header or []):
                    ws.append([date_str, time_str, boat, station_id, species, float(length_cm), float(confidence)])
                else:
                    # Backward compatibility with older files without Station ID column
                    ws.append([date_str, time_str, boat, species, float(length_cm), float(confidence)])
                wb.save(self.file_path)
        except LoggingError:
            raise
        except Exception as e:
            raise LoggingError(f"Failed to log entry: {e}") from e

    def cancel_last(self) -> bool:
        try:
            self._ensure_workbook()
            from openpyxl import load_workbook  # type: ignore

            with self.lock:
                wb = load_workbook(self.file_path)
                ws = wb.active
                if ws.max_row <= 1:
                    return False
                ws.delete_rows(ws.max_row, 1)
                wb.save(self.file_path)
                return True
        except Exception as e:
            raise LoggingError(f"Failed to cancel last entry: {e}") from e
