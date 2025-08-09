from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

from logger import ExcelLogger


def test_logger_append_and_cancel():
    with TemporaryDirectory() as td:
        log_path = Path(td) / "test_logs.xlsx"
        xl = ExcelLogger(str(log_path))
        xl.log_entry("Sea Bass", 27.5, 0.95)
        xl.log_entry("Cod", 30.0, 0.90)

        wb = load_workbook(log_path)
        ws = wb.active
        assert ws.max_row == 3  # header + 2 rows

        ok = xl.cancel_last()
        assert ok
        wb = load_workbook(log_path)
        ws = wb.active
        assert ws.max_row == 2  # header + 1 row
